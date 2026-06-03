import os
import io
import tempfile
import smtplib
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import core_transfer
import core_hacchu

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-me')

# バージョン管理（Semantic Versioning）
VERSION = "1.7.0"

# ── BOX OAuth2 認証（トークン自動保存） ─────────────────────────────
_TOKEN_FILE = '/volume1/webapp/box_tokens.json'

def _store_tokens(access_token, refresh_token):
    import json as _json, re
    try:
        with open(_TOKEN_FILE, 'w') as f:
            _json.dump({'access_token': access_token, 'refresh_token': refresh_token}, f)
    except Exception:
        pass
    # .env も同時に更新してNAS再起動後も最新トークンを維持
    try:
        env_path = '/volume1/webapp/.env'
        with open(env_path, 'r') as f:
            env = f.read()
        env = re.sub(r'BOX_ACCESS_TOKEN=.*', f'BOX_ACCESS_TOKEN={access_token}', env)
        env = re.sub(r'BOX_REFRESH_TOKEN=.*', f'BOX_REFRESH_TOKEN={refresh_token}', env)
        with open(env_path, 'w') as f:
            f.write(env)
    except Exception:
        pass

def _get_box_client():
    from boxsdk import OAuth2, Client
    import json as _json
    if os.path.exists(_TOKEN_FILE):
        with open(_TOKEN_FILE) as f:
            tokens = _json.load(f)
        access_token = tokens.get('access_token', '')
        refresh_token = tokens.get('refresh_token', os.environ['BOX_REFRESH_TOKEN'])
    else:
        access_token = os.environ.get('BOX_ACCESS_TOKEN', '')
        refresh_token = os.environ['BOX_REFRESH_TOKEN']
    auth = OAuth2(
        client_id=os.environ['BOX_CLIENT_ID'],
        client_secret=os.environ['BOX_CLIENT_SECRET'],
        access_token=access_token,
        refresh_token=refresh_token,
        store_tokens=_store_tokens,
    )
    return Client(auth)


def _box_token_keeper():
    import time
    # 48時間ごとにBOXへアクセスしてリフレッシュトークンの60日期限をリセット
    while True:
        time.sleep(48 * 3600)
        try:
            _get_box_client().user().get()
        except Exception:
            pass

threading.Thread(target=_box_token_keeper, daemon=True).start()


def _list_subfolders(folder_id):
    """指定フォルダ直下のサブフォルダ一覧を返す。ルートフォルダ自身も先頭に含む。"""
    client = _get_box_client()
    folder = client.folder(folder_id).get()
    result = [{'id': folder_id, 'name': f'📁 {folder.name}（このフォルダ）'}]
    items = client.folder(folder_id).get_items(limit=200)
    for item in items:
        if item.type == 'folder':
            result.append({'id': item.id, 'name': f'📂 {item.name}'})
    return result


def _list_excel_files(folder_id):
    client = _get_box_client()
    items = client.folder(folder_id).get_items(limit=200)
    files = [
        {'id': item.id, 'name': item.name}
        for item in items
        if item.type == 'file' and item.name.lower().endswith(('.xlsx', '.xlsm'))
    ]
    return sorted(files, key=lambda x: x['name'])


def _download_to_temp(file_id, suffix='.xlsx'):
    client = _get_box_client()
    content = client.file(file_id).content()
    fd, path = tempfile.mkstemp(suffix=suffix)
    with os.fdopen(fd, 'wb') as f:
        f.write(content)
    return path


def _get_box_file_name(file_id):
    client = _get_box_client()
    return client.file(file_id).get().name


def _update_box_file(file_id, file_bytes):
    client = _get_box_client()
    client.file(file_id).update_contents_with_stream(io.BytesIO(file_bytes))


def _upload_box_file(folder_id, filename, file_bytes):
    client = _get_box_client()
    return client.folder(folder_id).upload_stream(io.BytesIO(file_bytes), filename).id


def _upload_or_update_box_file(folder_id, filename, file_bytes):
    """同名ファイルが存在すれば上書き更新、なければ新規アップロード"""
    client = _get_box_client()
    folder = client.folder(folder_id)
    for item in folder.get_items(limit=200):
        if item.type == 'file' and item.name == filename:
            client.file(item.id).update_contents_with_stream(io.BytesIO(file_bytes))
            return item.id
    return folder.upload_stream(io.BytesIO(file_bytes), filename).id


def _send_email(file_bytes, filename, store_name, date_str):
    gmail_user = 'tomo.dream69@gmail.com'
    gmail_password = os.environ['GMAIL_APP_PASSWORD']
    gmail_to = os.environ['GMAIL_TO']

    subject = f'納品書 {date_str} {store_name}'
    msg = MIMEMultipart()
    msg['From'] = gmail_user
    msg['To'] = gmail_to
    msg['Subject'] = subject
    msg.attach(MIMEText(f'納品書を送付します。\n\n日付: {date_str}\n店舗: {store_name}', 'plain', 'utf-8'))

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(file_bytes)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(gmail_user, gmail_password)
        smtp.send_message(msg)


def _cleanup(*paths):
    for p in paths:
        try:
            if p and os.path.exists(p):
                os.unlink(p)
        except Exception:
            pass


def _add_notion_delivery_record(store_name, date_str, item_count):
    """Notion TODOページに納品書完了テキストを段落ブロックとして追記する。失敗しても例外を出さない。"""
    api_key = os.environ.get('NOTION_API_KEY', '')
    if not api_key:
        return
    date_display = f'{date_str[:2]}/{date_str[2:]}' if len(date_str) == 4 else date_str
    text = f'✅ 納品書作成済 {date_display} {store_name}（{item_count}件）'
    try:
        import requests as _req
        _req.patch(
            'https://api.notion.com/v1/blocks/36670bbc-e793-81bc-89a4-c64ad3c8e1ad/children',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Notion-Version': '2022-06-28',
                'Content-Type': 'application/json',
            },
            json={'children': [{'type': 'paragraph', 'paragraph': {
                'rich_text': [{'type': 'text', 'text': {'content': text}}]
            }}]},
            timeout=15,
        )
    except Exception as e:
        print(f'[Notion] {e}')


# ── 認証 ───────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == os.environ.get('APP_PASSWORD', 'inaho'):
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = 'パスワードが違います'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── トップ ─────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return render_template('index.html', version=VERSION)


# ── 通常モード ─────────────────────────────────────────────────
@app.route('/normal')
@login_required
def normal():
    order_folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')
    delivery_folder_id = os.environ.get('BOX_DELIVERY_FOLDER_ID', '')
    preselected_order_id = session.pop('fax_order_file_id', None)
    try:
        order_files = _list_excel_files(order_folder_id) if order_folder_id else []
        delivery_files = _list_excel_files(delivery_folder_id) if delivery_folder_id else []
    except Exception as e:
        return render_template('normal.html', error=f'BOX接続エラー: {e}',
                               order_files=[], delivery_files=[],
                               preselected_order_id=None)
    return render_template('normal.html', order_files=order_files,
                           delivery_files=delivery_files, error=None,
                           preselected_order_id=preselected_order_id)


@app.route('/api/stores', methods=['POST'])
@login_required
def api_stores():
    order_file_id = request.json.get('order_file_id')
    if not order_file_id:
        return jsonify({'error': '発注書が未選択'}), 400
    tmp = None
    try:
        tmp = _download_to_temp(order_file_id, suffix='.xlsx')
        stores = core_transfer.get_store_names(tmp)
        return jsonify({'stores': stores})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _cleanup(tmp)


@app.route('/normal/create', methods=['POST'])
@login_required
def normal_create():
    order_file_id = request.form.get('order_file_id')
    delivery_file_id = request.form.get('delivery_file_id')
    store_name = request.form.get('store_name', '').strip()
    date_str = request.form.get('date_str', '').replace('/', '').replace('-', '')

    if not all([order_file_id, delivery_file_id, store_name, date_str]):
        return render_template('normal.html', error='入力が不足しています',
                               order_files=[], delivery_files=[])

    order_tmp = delivery_tmp = None
    try:
        order_tmp = _download_to_temp(order_file_id, suffix='.xlsx')
        delivery_tmp = _download_to_temp(delivery_file_id, suffix='.xlsx')

        price_map = core_transfer.build_kubun_price_map(delivery_tmp)
        items = core_transfer.extract_order_items(order_tmp, store_name, price_map)

        if not items:
            return render_template('normal.html',
                                   error=f'「{store_name}」の商品データが見つかりませんでした',
                                   order_files=[], delivery_files=[])

        delivery_filename = _get_box_file_name(delivery_file_id)
        session['preview'] = {
            'delivery_file_id': delivery_file_id,
            'store_name': store_name,
            'date_str': date_str,
            'items': items,
            'delivery_filename': delivery_filename,
            'price_map': price_map,
        }
        return redirect(url_for('confirm'))

    except Exception as e:
        return render_template('normal.html', error=f'エラー: {e}',
                               order_files=[], delivery_files=[])
    finally:
        _cleanup(order_tmp, delivery_tmp)


# ── 手入力モード ───────────────────────────────────────────────
@app.route('/manual')
@login_required
def manual():
    delivery_folder_id = os.environ.get('BOX_DELIVERY_FOLDER_ID', '')
    try:
        delivery_files = _list_excel_files(delivery_folder_id) if delivery_folder_id else []
    except Exception as e:
        return render_template('manual.html', error=f'BOX接続エラー: {e}',
                               delivery_files=[])
    return render_template('manual.html', delivery_files=delivery_files, error=None)


@app.route('/manual/create', methods=['POST'])
@login_required
def manual_create():
    delivery_file_id = request.form.get('delivery_file_id')
    date_str = request.form.get('date_str', '').replace('/', '').replace('-', '')

    kubuns  = request.form.getlist('kubun[]')
    codes   = request.form.getlist('code[]')
    names   = request.form.getlist('name[]')
    hinbans = request.form.getlist('hinban[]')
    qtys    = request.form.getlist('qty[]')
    prices  = request.form.getlist('price[]')

    items = []
    for i in range(len(codes)):
        try:
            qty = int(qtys[i]) if qtys[i] else 0
            price = int(prices[i]) if prices[i] else 0
            if qty > 0 and codes[i]:
                items.append({
                    'kubun':  kubuns[i] if i < len(kubuns) else '',
                    'code':   codes[i],
                    'name':   names[i] if i < len(names) else '',
                    'hinban': hinbans[i] if i < len(hinbans) else '',
                    'qty':    qty,
                    'price':  price,
                    'amount': qty * price,
                })
        except (ValueError, IndexError):
            continue

    if not all([delivery_file_id, date_str]) or not items:
        return render_template('manual.html', error='入力が不足しています（商品は1行以上必要です）',
                               delivery_files=[])

    try:
        delivery_filename = _get_box_file_name(delivery_file_id)
        session['preview'] = {
            'delivery_file_id': delivery_file_id,
            'store_name': '手入力',
            'date_str': date_str,
            'items': items,
            'delivery_filename': delivery_filename,
            'price_map': {},
        }
        return redirect(url_for('confirm'))
    except Exception as e:
        return render_template('manual.html', error=f'エラー: {e}',
                               delivery_files=[])


# ── 確認・編集画面 ─────────────────────────────────────────────

@app.route('/confirm')
@login_required
def confirm():
    preview = session.get('preview')
    if not preview:
        return redirect(url_for('index'))
    return render_template('confirm.html', preview=preview, error=None)


@app.route('/confirm/pdf')
@login_required
def confirm_pdf():
    preview = session.get('preview')
    if not preview:
        return redirect(url_for('index'))
    try:
        pdf = core_transfer.generate_delivery_pdf(
            preview['items'], preview['store_name'], preview['date_str'])
        from flask import Response
        return Response(pdf, mimetype='application/pdf',
                        headers={'Content-Disposition': 'inline; filename="delivery.pdf"'})
    except Exception as e:
        return f'PDFプレビューエラー: {e}', 500


@app.route('/confirm/download')
@login_required
def confirm_download():
    preview = session.get('preview')
    if not preview:
        return redirect(url_for('index'))
    delivery_tmp = output_tmp = None
    try:
        delivery_tmp = _download_to_temp(preview['delivery_file_id'], suffix='.xlsx')
        fd, output_tmp = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        sheet_name, wb = core_transfer.add_delivery_sheet(
            delivery_tmp, preview['items'], preview['date_str'],
            output_tmp, preview.get('price_map') or {})
        wb.save(output_tmp)
        with open(output_tmp, 'rb') as f:
            file_bytes = f.read()
        from flask import send_file
        return send_file(
            io.BytesIO(file_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=preview['delivery_filename'],
        )
    except Exception as e:
        return f'ダウンロードエラー: {e}', 500
    finally:
        _cleanup(delivery_tmp, output_tmp)


@app.route('/confirm/upload', methods=['POST'])
@login_required
def confirm_upload():
    preview = session.get('preview')
    if not preview:
        return redirect(url_for('index'))
    uploaded = request.files.get('excel_file')
    if not uploaded:
        return render_template('confirm.html', preview=preview, error='ファイルが選択されていません')
    file_bytes = uploaded.read()
    try:
        _update_box_file(preview['delivery_file_id'], file_bytes)
        _send_email(file_bytes, preview['delivery_filename'],
                    preview['store_name'], preview['date_str'])
        threading.Thread(
            target=_add_notion_delivery_record,
            args=(preview['store_name'], preview['date_str'], len(preview['items'])),
            daemon=True,
        ).start()
        session.pop('preview', None)
        return render_template('success.html',
                               filename=preview['delivery_filename'],
                               store=preview['store_name'],
                               date_str=preview['date_str'],
                               item_count=len(preview['items']),
                               notion_sent=True)
    except Exception as e:
        return render_template('confirm.html', preview=preview, error=f'保存エラー: {e}')


@app.route('/confirm/save', methods=['POST'])
@login_required
def confirm_save():
    preview = session.get('preview')
    if not preview:
        return redirect(url_for('index'))

    kubuns  = request.form.getlist('kubun[]')
    codes   = request.form.getlist('code[]')
    names   = request.form.getlist('name[]')
    hinbans = request.form.getlist('hinban[]')
    qtys    = request.form.getlist('qty[]')
    prices  = request.form.getlist('price[]')

    items = []
    for i in range(len(codes)):
        try:
            qty = int(qtys[i]) if qtys[i] else 0
            price = int(prices[i]) if prices[i] else 0
            if codes[i]:
                items.append({
                    'kubun':  kubuns[i] if i < len(kubuns) else '',
                    'code':   codes[i],
                    'name':   names[i] if i < len(names) else '',
                    'hinban': hinbans[i] if i < len(hinbans) else '',
                    'qty':    qty,
                    'price':  price,
                    'amount': qty * price,
                })
        except (ValueError, IndexError):
            continue

    if not items:
        return render_template('confirm.html', preview=preview, error='商品が1行以上必要です')

    delivery_tmp = output_tmp = None
    try:
        delivery_tmp = _download_to_temp(preview['delivery_file_id'], suffix='.xlsx')
        fd, output_tmp = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        sheet_name, wb = core_transfer.add_delivery_sheet(
            delivery_tmp, items, preview['date_str'],
            output_tmp, preview.get('price_map') or {})
        wb.save(output_tmp)
        with open(output_tmp, 'rb') as f:
            file_bytes = f.read()
        _update_box_file(preview['delivery_file_id'], file_bytes)
        send_email = request.form.get('send_email', '0') == '1'
        if send_email:
            _send_email(file_bytes, preview['delivery_filename'],
                        preview['store_name'], preview['date_str'])
            threading.Thread(
                target=_add_notion_delivery_record,
                args=(preview['store_name'], preview['date_str'], len(items)),
                daemon=True,
            ).start()
        session.pop('preview', None)
        return render_template('success.html',
                               filename=preview['delivery_filename'],
                               store=preview['store_name'],
                               date_str=preview['date_str'],
                               item_count=len(items),
                               notion_sent=send_email)
    except Exception as e:
        preview['items'] = items
        return render_template('confirm.html', preview=preview, error=f'エラー: {e}')
    finally:
        _cleanup(delivery_tmp, output_tmp)


# ── 倉庫FAX確認 ────────────────────────────────────────────────

def _parse_fax_order(path):
    """発注書Excelの「稲穂」シートを解析して店舗・商品リストを返す。
    tab_fax.py の _fax_load_sheet() 相当。"""
    from openpyxl import load_workbook

    exclude = {'', 'None', '区分', '商品ｺｰﾄﾞ', '商品コード', '品名', '品番',
               'JANコード', 'JAN', '数量', '単価', '金額', '備考',
               'BB', '工具館', '店舗名', '引取'}
    PAGE_MARKS = {'①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩'}

    wb = load_workbook(path, data_only=True)
    result_stores = []
    result_items = []

    for ws in wb.worksheets:
        if '稲穂' not in ws.title:
            continue
        data = list(ws.iter_rows(values_only=True))

        store_col_map = {}
        store_row_vals = None
        for ri, row in enumerate(data[:15]):
            row_vals = [str(v or '').strip() for v in row]
            if '店舗名' in row_vals:
                candidates = [v for v in row_vals[4:] if v not in exclude]
                if candidates:
                    store_row_vals = row_vals
                else:
                    if ri + 1 < len(data):
                        store_row_vals = [str(v or '').strip() for v in data[ri + 1]]
                break

        if store_row_vals is None:
            continue

        for ci, v in enumerate(store_row_vals):
            if ci >= 4 and v not in exclude:
                store_col_map[v] = ci

        if not store_col_map:
            continue

        stores = list(store_col_map.keys())
        result_stores = stores
        items_map = {}

        for row in data:
            kubun = str(row[0] or '').strip()
            code = str(row[1] or '').strip()
            name = str(row[2] or '').replace('\n', '').strip()
            hinban = str(row[3] or '').strip()

            if kubun in PAGE_MARKS:
                continue
            if not code or code in ['商品ｺｰﾄﾞ', 'None', '']:
                continue
            if not kubun or kubun in ['区分', '企業名', '店舗名']:
                continue
            if not code.isdigit():
                continue

            if code not in items_map:
                items_map[code] = {
                    'code': code, 'name': name,
                    'kubun': kubun, 'hinban': hinban,
                    'qty_by_store': {}
                }

            for sname, ci in store_col_map.items():
                val = row[ci] if ci < len(row) else None
                try:
                    qty = int(val)
                except (TypeError, ValueError):
                    qty = 0
                existing = items_map[code]['qty_by_store'].get(sname, 0)
                items_map[code]['qty_by_store'][sname] = max(existing, qty)

        result_items = list(items_map.values())
        break  # 最初の「稲穂」シートのみ使用

    wb.close()
    return {'stores': result_stores, 'items': result_items}


def _write_fax_quantities(path, stores, quantities):
    """発注書Excelの数量を quantities で上書きして保存。
    quantities: {store_name: {code: qty}}"""
    from openpyxl import load_workbook

    exclude = {'', 'None', '区分', '商品ｺｰﾄﾞ', '商品コード', '品名', '品番',
               'JANコード', 'JAN', '数量', '単価', '金額', '備考',
               'BB', '工具館', '店舗名', '引取'}
    PAGE_MARKS = {'①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩'}

    wb = load_workbook(path)
    for ws in wb.worksheets:
        if '稲穂' not in ws.title:
            continue
        data = list(ws.iter_rows(values_only=False))

        store_col_map = {}
        store_row_vals = None
        for ri, row in enumerate(data[:15]):
            row_vals = [str(row[ci].value or '').strip() for ci in range(len(row))]
            if '店舗名' in row_vals:
                candidates = [v for v in row_vals[4:] if v not in exclude]
                if candidates:
                    store_row_vals = row_vals
                else:
                    if ri + 1 < len(data):
                        store_row_vals = [str(data[ri + 1][ci].value or '').strip()
                                          for ci in range(len(data[ri + 1]))]
                break

        if store_row_vals is None:
            continue

        for ci, v in enumerate(store_row_vals):
            if ci >= 4 and v not in exclude:
                store_col_map[v] = ci

        for row in data:
            kubun = str(row[0].value or '').strip()
            code = str(row[1].value or '').strip()
            if not code or code in ['商品ｺｰﾄﾞ', 'None']:
                continue
            if not kubun or kubun in ['区分', '企業名', '店舗名'] or kubun in PAGE_MARKS:
                continue
            for sname, ci in store_col_map.items():
                if sname in quantities and code in quantities[sname]:
                    qty = quantities[sname][code]
                    row[ci].value = qty if qty > 0 else None

    wb.save(path)
    wb.close()


@app.route('/fax')
@login_required
def fax():
    order_folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')
    try:
        order_files = _list_excel_files(order_folder_id) if order_folder_id else []
    except Exception as e:
        return render_template('fax.html', error=f'BOX接続エラー: {e}', order_files=[])
    return render_template('fax.html', order_files=order_files, error=None)


@app.route('/api/fax/table', methods=['POST'])
@login_required
def api_fax_table():
    order_file_id = request.json.get('order_file_id')
    if not order_file_id:
        return jsonify({'error': '発注書が未選択'}), 400
    tmp = None
    try:
        tmp = _download_to_temp(order_file_id, suffix='.xlsx')
        result = _parse_fax_order(tmp)
        if not result['stores']:
            return jsonify({'error': '「稲穂」シートまたは店舗情報が見つかりませんでした'}), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        _cleanup(tmp)


@app.route('/fax/save', methods=['POST'])
@login_required
def fax_save():
    import json as _json, shutil, tempfile as _tempfile
    order_file_id = request.form.get('order_file_id')
    stores_json = request.form.get('stores', '[]')
    try:
        stores = _json.loads(stores_json)
    except Exception:
        stores = []

    quantities = {}
    for key, val in request.form.items():
        if not key.startswith('qty_'):
            continue
        parts = key[4:].split('_', 1)
        if len(parts) != 2:
            continue
        code, store_idx_str = parts
        try:
            store_idx = int(store_idx_str)
            store_name = stores[store_idx]
            qty = int(val) if val else 0
            quantities.setdefault(store_name, {})[code] = qty
        except (ValueError, IndexError):
            continue

    tmp = output_tmp = None
    try:
        order_folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')
        tmp = _download_to_temp(order_file_id, suffix='.xlsx')
        fd, output_tmp = _tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        shutil.copy2(tmp, output_tmp)
        _write_fax_quantities(output_tmp, stores, quantities)
        with open(output_tmp, 'rb') as f:
            file_bytes = f.read()
        orig_filename = _get_box_file_name(order_file_id)
        base, ext = os.path.splitext(orig_filename)
        new_filename = base + '_数量確定' + ext
        new_file_id = _upload_or_update_box_file(order_folder_id, new_filename, file_bytes)
        session['fax_order_file_id'] = new_file_id
        return redirect(url_for('normal'))
    except Exception as e:
        order_folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')
        try:
            order_files = _list_excel_files(order_folder_id)
        except Exception:
            order_files = []
        return render_template('fax.html', error=f'保存エラー: {e}', order_files=order_files)
    finally:
        _cleanup(tmp, output_tmp)


# ── 発注書変換 ──────────────────────────────────────────────────

def _hacchu_folders():
    order_folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')
    try:
        return _list_subfolders(order_folder_id) if order_folder_id else []
    except Exception:
        return [{'id': order_folder_id, 'name': '📁 発注書フォルダ'}] if order_folder_id else []


@app.route('/hacchu', methods=['GET', 'POST'])
@login_required
def hacchu():
    if request.method == 'GET':
        folders = _hacchu_folders()
        return render_template('hacchu.html', error=None, saved=None,
                               folders=folders, version=VERSION)

    files = request.files.getlist('order_files')
    folder_id = request.form.get('folder_id', '').strip()

    if not files or all(f.filename == '' for f in files):
        folders = _hacchu_folders()
        return render_template('hacchu.html', error='ファイルを選択してください',
                               saved=None, folders=folders, version=VERSION)

    if not folder_id:
        folder_id = os.environ.get('BOX_ORDER_FOLDER_ID', '')

    tmp_paths = []
    try:
        for f in files:
            if not f.filename:
                continue
            fd, tmp = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            f.save(tmp)
            tmp_paths.append(tmp)

        if not tmp_paths:
            folders = _hacchu_folders()
            return render_template('hacchu.html', error='有効なファイルがありません',
                                   saved=None, folders=folders, version=VERSION)

        from datetime import date
        mmdd = f"{date.today().month:02d}{date.today().day:02d}"
        output_filename = f"稲穂_{mmdd}_発注書.xlsx"

        file_bytes = core_hacchu.web_merge_orders(tmp_paths)
        _upload_or_update_box_file(folder_id, output_filename, file_bytes)

        folders = _hacchu_folders()
        return render_template('hacchu.html', error=None, saved=output_filename,
                               folders=folders, selected_folder=folder_id, version=VERSION)
    except Exception as e:
        folders = _hacchu_folders()
        return render_template('hacchu.html', error=f'変換エラー: {e}',
                               saved=None, folders=folders, version=VERSION)
    finally:
        _cleanup(*tmp_paths)


if __name__ == '__main__':
    app.run(debug=True, port=5010, host='0.0.0.0', use_reloader=False)
