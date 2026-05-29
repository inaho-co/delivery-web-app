import os
import re
from datetime import datetime
from pathlib import Path

# ============================================================
# コアロジック
# ============================================================

def get_thin_border():
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin")
    return Border(top=thin, bottom=thin, left=thin, right=thin)

def get_border_for_col(col):
    """既存シートの列ごとの罫線スタイルを再現"""
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin")
    none = Side(border_style=None)
    if col == 1:
        return Border(top=thin, bottom=thin, left=thin, right=none)
    elif col in [2, 3]:
        return Border(top=thin, bottom=thin, left=none, right=none)
    elif col == 4:
        return Border(top=thin, bottom=thin, left=none, right=thin)
    else:
        return Border(top=thin, bottom=thin, left=thin, right=thin)

def build_kubun_price_map(delivery_path):
    """納品書の既存シートから区分ごとの単価を取得。
    通常フォーマット: {区分: 単価}
    税込フォーマット: {区分: {'tax_ex': 税抜単価, 'tax_in': 税込単価}}
    """
    kubun_price = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(delivery_path, data_only=True)
        for ws in wb.worksheets:
            header_vals = {cell.column: str(cell.value or '') for cell in ws[7]}
            is_tax = any('税込' in v for v in header_vals.values())

            for r in range(8, ws.max_row + 1):
                if is_tax:
                    kubun = str(ws.cell(r, 12).value or '').strip()
                    price_ex = ws.cell(r, 9).value
                    price_in = ws.cell(r, 10).value
                    if kubun and price_ex and price_in:
                        try:
                            if kubun not in kubun_price:
                                kubun_price[kubun] = {'tax_ex': int(price_ex), 'tax_in': int(price_in)}
                        except (TypeError, ValueError):
                            pass
                else:
                    kubun = str(ws.cell(r, 12).value or '').strip()
                    price = ws.cell(r, 10).value
                    if kubun and price:
                        try:
                            if kubun not in kubun_price:
                                kubun_price[kubun] = int(price)
                        except (TypeError, ValueError):
                            pass
            if len(kubun_price) >= 2:
                break
    except Exception:
        pass
    return kubun_price


def build_code_price_map(delivery_path):
    """納品書の全シートからコード→単価のマッピングを取得。
    返り値: {code_str: [単価, ...]}（重複なし、int のリスト）
    税込フォーマットの場合は税抜単価（col9）を使用。
    """
    code_price = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(delivery_path, data_only=True)
        for ws in wb.worksheets:
            header_vals = {cell.column: str(cell.value or '') for cell in ws[7]}
            is_tax = any('税込' in v for v in header_vals.values())

            # 税込: コード=col5, 税抜単価=col9 / 税抜: コード=col6, 単価=col10
            if is_tax:
                col_code_idx, col_price_idx = 5, 9
            else:
                col_code_idx, col_price_idx = 6, 10

            for r in range(8, ws.max_row + 1):
                code_val  = ws.cell(r, col_code_idx).value
                price_val = ws.cell(r, col_price_idx).value
                if not code_val or not price_val:
                    continue
                code_str = str(code_val).strip()
                if not code_str:
                    continue
                try:
                    price_int = int(float(price_val))
                except (TypeError, ValueError):
                    continue
                if price_int <= 0:
                    continue
                if code_str not in code_price:
                    code_price[code_str] = []
                if price_int not in code_price[code_str]:
                    code_price[code_str].append(price_int)
    except Exception:
        pass
    return code_price


def extract_order_items(order_path, store_name, price_map=None):
    """発注書から指定店舗の数量ある商品を抽出（シート名に「稲穂」を含むシートのみ）"""
    from openpyxl import load_workbook
    wb = load_workbook(order_path, data_only=True)
    items = []

    for ws in wb.worksheets:
        if '稲穂' not in ws.title:
            continue
        data = list(ws.iter_rows(values_only=True))

        store_col = None
        for r, row in enumerate(data[:10]):
            for c, val in enumerate(row):
                if str(val or '').strip() == store_name:
                    store_col = c
                    break
            if store_col is not None:
                break

        if store_col is None:
            continue

        for row in data:
            kubun = str(row[0] or '').strip()
            code  = str(row[1] or '').strip()
            name  = str(row[2] or '').replace('\n', '').strip()
            hinban = str(row[3] or '').strip()
            qty_val = row[store_col] if store_col < len(row) else None

            if not code or code in ['商品ｺｰﾄﾞ', 'None', '']:
                continue
            if not kubun or kubun in ['区分', '企業名', '店舗名'] or \
               any(kubun == c for c in ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩']):
                continue

            try:
                qty = int(qty_val)
            except (TypeError, ValueError):
                continue

            if qty <= 0:
                continue

            if price_map and kubun in price_map:
                p = price_map[kubun]
                price = p['tax_ex'] if isinstance(p, dict) else p
            elif price_map and any('1000' in k for k in price_map) and '1000' in kubun:
                p = next(v for k, v in price_map.items() if '1000' in k)
                price = p['tax_ex'] if isinstance(p, dict) else p
            elif price_map and any('600' in k for k in price_map) and '600' in kubun:
                p = next(v for k, v in price_map.items() if '600' in k)
                price = p['tax_ex'] if isinstance(p, dict) else p
            elif '1000' in kubun:
                price = 730
            elif '600' in kubun:
                price = 420
            elif code in ['100901', '100902', '100903', '100904']:
                price = 220  # midoho EDLP商品（税込242円は納品書側で×1.1）
            else:
                price = 0

            items.append({
                'kubun': kubun,
                'code': code,
                'name': name,
                'hinban': hinban,
                'qty': qty,
                'price': price,
                'amount': qty * price,
            })

    return items


def get_store_names(order_path):
    """発注書から店舗名一覧を取得（シート名に「稲穂」を含むシートのみ対象）。"""
    from openpyxl import load_workbook
    wb = load_workbook(order_path, data_only=True)
    stores = []

    exclude = {'', 'None', '区分', '商品ｺｰﾄﾞ', '商品コード', '品名', '品番',
               'JANコード', 'JAN', '数量', '単価', '金額', '備考',
               'BB', '工具館', '店舗名', '引取'}

    for ws in wb.worksheets:
        if '稲穂' not in ws.title:
            continue

        rows = list(ws.iter_rows(max_row=10, values_only=True))
        store_row_vals = None

        for row in rows:
            row_vals = [str(v or '').strip() for v in row]
            if '店舗名' in row_vals:
                candidates = [v for v in row_vals[4:] if v not in exclude]
                if candidates:
                    store_row_vals = row_vals
                    break
                else:
                    idx = rows.index(row)
                    if idx + 1 < len(rows):
                        store_row_vals = [str(v or '').strip() for v in rows[idx + 1]]
                    break

        if store_row_vals:
            for v in store_row_vals[4:]:
                if v not in exclude and v not in stores:
                    stores.append(v)

    return stores


def copy_cell_style(src, dst):
    """セルのスタイルを完全コピー（ハイパーリンク・数式は絶対にコピーしない）"""
    if src.has_style:
        dst.font      = src.font.copy()
        dst.border    = src.border.copy()
        dst.fill      = src.fill.copy()
        dst.alignment = src.alignment.copy()
        dst.number_format = src.number_format
    dst.hyperlink = None
    # 数式が紛れ込んでいた場合は除去（Excelの警告防止）
    if isinstance(dst.value, str) and dst.value.startswith('='):
        dst.value = None

def copy_row_style(src_ws, src_row, dst_ws, dst_row, max_col=12):
    """指定行のスタイルだけ別の行にコピー（値・数式はコピーしない）"""
    for col in range(1, max_col + 1):
        copy_cell_style(src_ws.cell(row=src_row, column=col),
                        dst_ws.cell(row=dst_row, column=col))
        # コピー先の値を必ずクリア（数式の混入防止）
        dst_ws.cell(row=dst_row, column=col).value = None


def remove_mailto_from_file(path):
    """xlsxファイルからmailtoハイパーリンクをXMLレベルで完全除去して上書き保存
    【修正】原子的な置き換えで破損リスクを低減
    """
    import zipfile, re, tempfile, shutil

    # mailtoが含まれているか先にチェック
    has_mailto = False
    try:
        with zipfile.ZipFile(path, 'r') as z:
            for name in z.namelist():
                try:
                    if 'mailto' in z.read(name).decode('utf-8'):
                        has_mailto = True
                        break
                except Exception:
                    pass
    except Exception:
        return False

    if not has_mailto:
        return False

    # 元ファイルと同じフォルダに一時ファイルを作成（同一ドライブ保証）
    dir_name = os.path.dirname(os.path.abspath(path))
    tmp_fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=dir_name)
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(path, 'r') as zin:
            with zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    try:
                        text = data.decode('utf-8')
                        text = re.sub(r'<hyperlinks>.*?</hyperlinks>', '', text, flags=re.DOTALL)
                        text = re.sub(r'<Relationship[^>]*hyperlink[^>]*/>', '', text)
                        data = text.encode('utf-8')
                    except Exception:
                        pass
                    zout.writestr(item, data)

        # 【修正】書き込み完了後にバックアップ→置き換え（原子的）
        backup = path + '.bak'
        shutil.copy2(path, backup)   # 元ファイルをバックアップ
        shutil.move(tmp, path)       # 新ファイルで上書き（moveは原子的）
        try:
            os.remove(backup)        # バックアップ削除（失敗しても無視）
        except Exception:
            pass
        return True

    except Exception:
        # 失敗した場合は一時ファイルだけ削除（元ファイルは触らない）
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return False


def strip_all_hyperlinks(wb):
    """ワークブック内の全シート・全セルのハイパーリンクを完全削除"""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    cell.hyperlink = None
        if hasattr(ws, '_hyperlinks'):
            ws._hyperlinks = []


def add_delivery_sheet(delivery_path, items, date_str, output_path, price_map=None, wb=None):
    """
    納品書に新シートを追加。
    既存シートは一切変更しない。最後のシートのレイアウトを完全再現して新シートを作成。
    wb を渡した場合はそのワークブックを使い回す（複数店舗対応）。

    【修正】wb が渡された場合（複数店舗処理）は save() しない。
    　　　　呼び出し元で最後に1回だけ save() すること。
    """
    caller_owns_wb = (wb is not None)  # 呼び出し元がwbを管理しているか

    if wb is None:
        try:
            remove_mailto_from_file(delivery_path)
        except Exception:
            pass
        from openpyxl import load_workbook
        wb = load_workbook(delivery_path)
        strip_all_hyperlinks(wb)

    tmpl = next(
    (ws for ws in reversed(wb.worksheets) if '封筒' not in ws.title),
    wb.worksheets[-1]
)

    month = int(date_str[:2])
    day   = int(date_str[2:]) if len(date_str) > 2 else None

    if len(date_str) == 2:
        sheet_name = date_str + '--'
    else:
        sheet_name = date_str
    if sheet_name in wb.sheetnames:
        base = sheet_name
        i = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f'{base}_{i}'
            i += 1

    new_ws = wb.create_sheet(title=sheet_name)

    HEADER_ROWS = 7
    for r in range(1, HEADER_ROWS + 1):
        for src_cell in tmpl[r]:
            dst_cell = new_ws.cell(row=r, column=src_cell.column)
            # 数式はコピーしない（新シートで参照先がずれて壊れるのを防ぐ）
            v = src_cell.value
            if isinstance(v, str) and v.startswith('='):
                dst_cell.value = None
            else:
                dst_cell.value = v
            copy_cell_style(src_cell, dst_cell)
        new_ws.row_dimensions[r].height = tmpl.row_dimensions[r].height if r in tmpl.row_dimensions else 18

    # 【修正】結合セルはリスト化してから処理（イテレート中の変更による破損を防ぐ）
    merges_to_copy = [str(m) for m in tmpl.merged_cells.ranges if m.min_row <= HEADER_ROWS]
    for merge_str in merges_to_copy:
        try:
            new_ws.merge_cells(merge_str)
        except Exception:
            pass

    # 列幅コピー：登録済み列はそのままコピー
    for col_letter, dim in tmpl.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width

    # A〜D列：テンプレートに登録がない列はA列の幅を使う（日付部分の幅ズレ防止）
    a_width = tmpl.column_dimensions['A'].width if 'A' in tmpl.column_dimensions else 3.625
    for col_letter in ['A', 'B', 'C', 'D']:
        if col_letter not in tmpl.column_dimensions or not tmpl.column_dimensions[col_letter].width:
            new_ws.column_dimensions[col_letter].width = a_width

    # J・K列：税込納品書で未登録になりやすいため、I列の幅で補完（幅狭くなる防止）
    for col_letter, fallback in [('J', 'I'), ('K', 'I')]:
        if col_letter not in tmpl.column_dimensions or not tmpl.column_dimensions[col_letter].width:
            fb_width = (tmpl.column_dimensions[fallback].width
                        if fallback in tmpl.column_dimensions
                        else 8.43)
            new_ws.column_dimensions[col_letter].width = fb_width

    STYLE_REF_ROW = HEADER_ROWS + 1

    header_row_vals = {cell.column: str(cell.value or '') for cell in tmpl[HEADER_ROWS]}
    is_tax_included = any('税込' in v for v in header_row_vals.values())

    if is_tax_included:
        col_code   = 5
        col_name   = 6
        col_hinban = 7
        col_qty    = 8
        col_price  = 9
        col_price2 = 10
        col_amount = 11
        col_kubun  = 12
    else:
        col_code   = 6
        col_name   = 7
        col_hinban = 8
        col_qty    = 9
        col_price  = 10
        col_price2 = None
        col_amount = 11
        col_kubun  = 12

    page_total_tmpl_row = tmpl.max_row
    for r in range(tmpl.max_row, 0, -1):
        for cell in tmpl[r]:
            if str(cell.value or '') == '頁計':
                page_total_tmpl_row = r
                break
        else:
            continue
        break

    write_row = HEADER_ROWS + 1
    prev_kubun = None

    for item in items:
        if prev_kubun and item['kubun'] != prev_kubun:
            copy_row_style(tmpl, STYLE_REF_ROW, new_ws, write_row)
            cell = new_ws.cell(row=write_row, column=col_amount)
            cell.value = 0
            cell.number_format = '#,##0;-#,##0;"-"'
            new_ws.row_dimensions[write_row].height = 18
            write_row += 1

        copy_row_style(tmpl, STYLE_REF_ROW, new_ws, write_row)
        new_ws.row_dimensions[write_row].height = 18

        new_ws.cell(row=write_row, column=1).value         = month
        new_ws.cell(row=write_row, column=2).value         = '月'
        new_ws.cell(row=write_row, column=3).value         = day if day is not None else None
        new_ws.cell(row=write_row, column=4).value         = '日'
        new_ws.cell(row=write_row, column=col_code).value  = item['code']
        new_ws.cell(row=write_row, column=col_name).value  = item['name']
        new_ws.cell(row=write_row, column=col_hinban).value = item['hinban']
        new_ws.cell(row=write_row, column=col_qty).value   = item['qty']
        new_ws.cell(row=write_row, column=col_price).value = item['price']
        if col_price2:
            price_in = None
            if price_map:
                kubun = item['kubun']
                p = price_map.get(kubun)
                if p is None:
                    p = next((v for k, v in price_map.items() if '1000' in k and '1000' in kubun), None)
                if p is None:
                    p = next((v for k, v in price_map.items() if '600' in k and '600' in kubun), None)
                if isinstance(p, dict):
                    price_in = p.get('tax_in')
            if price_in is None:
                price_in = int(item['price'] * 1.1)
            new_ws.cell(row=write_row, column=col_price2).value = price_in
        new_ws.cell(row=write_row, column=col_amount).value = \
            f'=H{write_row}*J{write_row}' if is_tax_included else f'=I{write_row}*J{write_row}'
        new_ws.cell(row=write_row, column=col_kubun).value = item['kubun']

        prev_kubun = item['kubun']
        write_row += 1

    for src_cell in tmpl[page_total_tmpl_row]:
        copy_cell_style(src_cell, new_ws.cell(row=write_row, column=src_cell.column))
    new_ws.row_dimensions[write_row].height = 18
    page_label_col = col_price2 if col_price2 else col_price
    new_ws.cell(row=write_row, column=page_label_col).value = '頁計'
    new_ws.cell(row=write_row, column=col_amount).value = \
        f'=SUM(K{HEADER_ROWS+1}:K{write_row-1})'

    total_rows = max(write_row, 200)
    for r in range(1, total_rows + 1):
        new_ws.row_dimensions[r].height = 18
    new_ws.sheet_format.defaultRowHeight = 18
    new_ws.sheet_format.customHeight = True
    new_ws.sheet_view.view      = 'pageBreakPreview'
    new_ws.sheet_view.zoomScale = 100
    new_ws.page_setup.scale = 90
    new_ws.page_margins.top    = 1 / 2.54
    new_ws.page_margins.bottom = 0
    new_ws.page_margins.left   = 0
    new_ws.page_margins.right  = 0
    new_ws.page_margins.header = 0
    new_ws.page_margins.footer = 0
    new_ws.print_options.horizontalCentered = True
    new_ws.print_options.verticalCentered   = False
    new_ws.oddHeader.center.text = '&P/&N ページ'
    new_ws.oddHeader.center.size = 9
    # print_title_rows（タイトル行 1:7）を設定
    # openpyxl の属性で設定し、重複definedNamesによる破損を防ぐため
    # 保存前に run_process 側で remove_external_links を呼ぶことで対処
    try:
        new_ws.print_title_rows = '1:7'
    except Exception:
        pass

    # 【修正】wb が呼び出し元から渡された場合はここでは save() しない。
    # 　　　　単独呼び出し（wb=None）の場合のみここで保存する。
    if not caller_owns_wb:
        wb.save(output_path)

    return sheet_name, wb


def add_invoice_details(invoice_path, items, date_str, output_path, delivery_path=None):
    """
    請求書の月末シートの頁計の上に明細を挿入。
    スタイルは納品書の既存シートから参照する。
    """
    from openpyxl import load_workbook
    wb = load_workbook(invoice_path)
    strip_all_hyperlinks(wb)

    month = int(date_str[:2])
    day   = int(date_str[2:])
    target_name = f'{month}月末'

    style_ref_ws = None
    style_ref_row = 8
    if delivery_path:
        try:
            from openpyxl import load_workbook
            wb_delivery = load_workbook(delivery_path)
            strip_all_hyperlinks(wb_delivery)
            style_ref_ws = wb_delivery.worksheets[-1]
        except Exception:
            style_ref_ws = None

    if target_name not in wb.sheetnames:
        last_end = next((n for n in reversed(wb.sheetnames) if '月末' in n), None)
        if not last_end:
            raise ValueError('月末シートが見つかりません')
        src_ws = wb[last_end]

        new_ws = wb.create_sheet(title=target_name)

        for src_cell in src_ws[1]:
            dst_cell = new_ws.cell(row=1, column=src_cell.column)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)
        new_ws.row_dimensions[1].height = src_ws.row_dimensions[1].height if 1 in src_ws.row_dimensions else 18

        for col_letter, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width

        first_page_row = None
        for r in range(2, src_ws.max_row + 1):
            for cell in src_ws[r]:
                if str(cell.value or '').strip() == '頁計':
                    first_page_row = r
                    break
            if first_page_row:
                break
        if first_page_row:
            for src_cell in src_ws[first_page_row]:
                dst_cell = new_ws.cell(row=2, column=src_cell.column)
                copy_cell_style(src_cell, dst_cell)
            new_ws.cell(row=2, column=10).value = '頁計'

    ws = wb[target_name]

    last_data_row = 1
    for r in range(ws.max_row, 0, -1):
        has_data = False
        for cell in ws[r]:
            v = str(cell.value or '').strip()
            if v and v not in ['頁計', '前頁より']:
                has_data = True
                break
        if has_data:
            last_data_row = r
            break

    insert_before = last_data_row + 1
    insert_count = len(items) + 1

    for r in range(ws.max_row, insert_before - 1, -1):
        for col in range(1, 13):
            src_cell = ws.cell(row=r, column=col)
            dst_cell = ws.cell(row=r + insert_count, column=col)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)
        ws.row_dimensions[r + insert_count].height = ws.row_dimensions[r].height if r in ws.row_dimensions else 18

    from openpyxl.styles import Border
    no_border = Border()
    for r in range(insert_before, insert_before + insert_count):
        for col in range(1, 13):
            cell = ws.cell(row=r, column=col)
            cell.value = None
            cell.border = no_border

    if style_ref_ws is None:
        style_ref_ws = ws
        style_ref_row = 2
        for r in range(2, min(15, ws.max_row + 1)):
            try:
                if ws.cell(row=r, column=1).value and int(ws.cell(row=r, column=1).value) in range(1, 13):
                    style_ref_row = r
                    break
            except (TypeError, ValueError):
                pass

    write_row = insert_before
    prev_kubun = None

    for item in items:
        if prev_kubun and item['kubun'] != prev_kubun:
            copy_row_style(style_ref_ws, style_ref_row, ws, write_row)
            ws.cell(row=write_row, column=11).value = 0
            ws.row_dimensions[write_row].height = 18
            write_row += 1

        copy_row_style(style_ref_ws, style_ref_row, ws, write_row)
        ws.row_dimensions[write_row].height = 18

        ws.cell(row=write_row, column=1).value  = month
        ws.cell(row=write_row, column=2).value  = '月'
        ws.cell(row=write_row, column=3).value  = day
        ws.cell(row=write_row, column=4).value  = '日'
        ws.cell(row=write_row, column=6).value  = item['code']
        ws.cell(row=write_row, column=7).value  = item['name']
        ws.cell(row=write_row, column=8).value  = item['hinban']
        ws.cell(row=write_row, column=9).value  = item['qty']
        ws.cell(row=write_row, column=10).value = item['price']
        ws.cell(row=write_row, column=11).value = f'=I{write_row}*J{write_row}'
        ws.cell(row=write_row, column=12).value = item['kubun']

        prev_kubun = item['kubun']
        write_row += 1

    return wb


def add_kattorikoku_delivery_sheet(delivery_path, date_str, qty_1000, qty_600, wb=None):
    """買取王国用納品書シートを追加（1000円均一・600円均一の2行フォーマット）"""
    caller_owns_wb = (wb is not None)
    if wb is None:
        try:
            remove_mailto_from_file(delivery_path)
        except Exception:
            pass
        from openpyxl import load_workbook
        wb = load_workbook(delivery_path)
        strip_all_hyperlinks(wb)

    # 封筒シートを除いた最後のシートをテンプレートとして使用
    tmpl = next(
        (ws for ws in reversed(wb.worksheets) if '封筒' not in ws.title),
        wb.worksheets[-1]
    )

    month = int(date_str[:2])
    day   = int(date_str[2:]) if len(date_str) > 2 else None

    sheet_name = '--' if len(date_str) == 2 else date_str
    if sheet_name in wb.sheetnames:
        base = sheet_name
        i = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f'{base}_{i}'
            i += 1

    new_ws = wb.create_sheet(title=sheet_name)

    # ヘッダー行（1〜7行）をテンプレートからコピー
    HEADER_ROWS = 7
    for r in range(1, HEADER_ROWS + 1):
        for src_cell in tmpl[r]:
            dst_cell = new_ws.cell(row=r, column=src_cell.column)
            v = src_cell.value
            if isinstance(v, str) and v.startswith('='):
                dst_cell.value = None
            else:
                dst_cell.value = v
            copy_cell_style(src_cell, dst_cell)
        new_ws.row_dimensions[r].height = (
            tmpl.row_dimensions[r].height if r in tmpl.row_dimensions else 18)

    # 結合セルのコピー
    merges_to_copy = [str(m) for m in tmpl.merged_cells.ranges if m.min_row <= HEADER_ROWS]
    for merge_str in merges_to_copy:
        try:
            new_ws.merge_cells(merge_str)
        except Exception:
            pass

    # 列幅のコピー・補完
    for col_letter, dim in tmpl.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width
    a_width = tmpl.column_dimensions['A'].width if 'A' in tmpl.column_dimensions else 3.625
    for col_letter in ['A', 'B', 'C', 'D']:
        if col_letter not in tmpl.column_dimensions or not tmpl.column_dimensions[col_letter].width:
            new_ws.column_dimensions[col_letter].width = a_width
    for col_letter, fallback in [('J', 'I'), ('K', 'I')]:
        if col_letter not in tmpl.column_dimensions or not tmpl.column_dimensions[col_letter].width:
            fb_width = (tmpl.column_dimensions[fallback].width
                        if fallback in tmpl.column_dimensions else 8.43)
            new_ws.column_dimensions[col_letter].width = fb_width

    STYLE_REF_ROW = HEADER_ROWS + 1

    # 書き込む商品（数量0はスキップ）
    items_to_write = []
    if qty_1000 > 0:
        items_to_write.append({'code': '001000', 'name': '1000円均一',
                                'qty': qty_1000, 'price_ex': 700, 'price_in': 770})
    if qty_600 > 0:
        items_to_write.append({'code': '000600', 'name': '600円均一',
                                'qty': qty_600, 'price_ex': 390, 'price_in': 429})

    write_row = HEADER_ROWS + 1
    for item in items_to_write:
        copy_row_style(tmpl, STYLE_REF_ROW, new_ws, write_row)
        new_ws.row_dimensions[write_row].height = 18
        new_ws.cell(row=write_row, column=1).value  = month
        new_ws.cell(row=write_row, column=2).value  = '月'
        new_ws.cell(row=write_row, column=3).value  = day
        new_ws.cell(row=write_row, column=4).value  = '日'
        new_ws.cell(row=write_row, column=5).value  = item['code']      # 商品コード
        new_ws.cell(row=write_row, column=6).value  = item['name']      # 品名
        new_ws.cell(row=write_row, column=7).value  = None              # 品番
        new_ws.cell(row=write_row, column=8).value  = item['qty']       # 数量
        new_ws.cell(row=write_row, column=9).value  = item['price_ex']  # 単価(税抜)
        new_ws.cell(row=write_row, column=10).value = item['price_in']  # 単価(税込)
        new_ws.cell(row=write_row, column=11).value = f'=H{write_row}*J{write_row}'  # 金額(税込)
        new_ws.cell(row=write_row, column=12).value = None              # 備考
        write_row += 1

    # 頁計行（テンプレートの頁計行スタイルを流用）
    page_total_tmpl_row = tmpl.max_row
    for r in range(tmpl.max_row, 0, -1):
        for cell in tmpl[r]:
            if str(cell.value or '') == '頁計':
                page_total_tmpl_row = r
                break
        else:
            continue
        break
    for src_cell in tmpl[page_total_tmpl_row]:
        copy_cell_style(src_cell, new_ws.cell(row=write_row, column=src_cell.column))
    new_ws.row_dimensions[write_row].height = 18
    new_ws.cell(row=write_row, column=10).value = '頁計'
    new_ws.cell(row=write_row, column=11).value = f'=SUM(K{HEADER_ROWS+1}:K{write_row-1})'

    # ページ設定
    total_rows = max(write_row, 200)
    for r in range(1, total_rows + 1):
        new_ws.row_dimensions[r].height = 18
    new_ws.sheet_format.defaultRowHeight = 18
    new_ws.sheet_format.customHeight = True
    new_ws.sheet_view.view      = 'pageBreakPreview'
    new_ws.sheet_view.zoomScale = 100
    new_ws.page_setup.scale = 90
    new_ws.page_margins.top    = 1 / 2.54
    new_ws.page_margins.bottom = 0
    new_ws.page_margins.left   = 0
    new_ws.page_margins.right  = 0
    new_ws.page_margins.header = 0
    new_ws.page_margins.footer = 0
    new_ws.print_options.horizontalCentered = True
    new_ws.print_options.verticalCentered   = False
    new_ws.oddHeader.center.text = '&P/&N ページ'
    new_ws.oddHeader.center.size = 9
    try:
        new_ws.print_title_rows = '1:7'
    except Exception:
        pass

    if not caller_owns_wb:
        wb.save(delivery_path)

    return sheet_name, wb


def generate_delivery_pdf(items, store_name, date_str):
    """納品書PDFを生成してbytesを返す。reportlab + fonts/IPAexGothic.ttf を使用。"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os

    font_path = _os.path.join(_os.path.dirname(__file__), 'fonts', 'IPAexGothic.ttf')
    pdfmetrics.registerFont(TTFont('IPA', font_path))
    F = 'IPA'

    date_display = f'{date_str[:2]}/{date_str[2:]}' if len(date_str) == 4 else date_str
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    story = [
        Paragraph('納　品　書', ParagraphStyle('t', fontName=F, fontSize=16, alignment=1, spaceAfter=6)),
        Spacer(1, 4*mm),
        Paragraph(f'店舗：{store_name}　日付：{date_display}',
                  ParagraphStyle('i', fontName=F, fontSize=11, spaceAfter=4)),
        Spacer(1, 4*mm),
    ]
    col_w = [18*mm, 22*mm, 55*mm, 28*mm, 14*mm, 18*mm, 22*mm]
    rows = [['区分', 'コード', '品名', '品番', '数量', '単価', '金額']]
    total = 0
    for item in items:
        amt = item.get('amount', item['qty'] * item['price'])
        total += amt
        rows.append([item.get('kubun', ''), item.get('code', ''), item.get('name', ''),
                     item.get('hinban', ''), str(item['qty']),
                     f"{item['price']:,}", f"{amt:,}"])
    rows.append(['', '', '', '', '', '合　計', f"{total:,}"])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME',       (0, 0), (-1, -1), F),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('BACKGROUND',     (0, 0), (-1,  0), colors.HexColor('#1a2332')),
        ('TEXTCOLOR',      (0, 0), (-1,  0), colors.white),
        ('ALIGN',          (4, 0), (-1, -1), 'RIGHT'),
        ('ALIGN',          (0, 0), ( 3, -1), 'LEFT'),
        ('GRID',           (0, 0), (-1, -2), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND',     (0, -1), (-1, -1), colors.HexColor('#f0f4f8')),
        ('LINEABOVE',      (0, -1), (-1, -1), 1.5, colors.HexColor('#1a2332')),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()
