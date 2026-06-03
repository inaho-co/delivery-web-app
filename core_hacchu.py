"""
発注書変換ロジック（スマホ版delivery-web-app用）
パソコン版 core_hacchu.py の主要関数 + web用ラッパーを含む
"""
import os
import re
import io
from datetime import date


# ── 共通ユーティリティ ──────────────────────────────────────────

def hacchu_get_store_cols(ws):
    """E列（5列目）以降で、row1に値がある列番号のリストを返す"""
    cols = []
    max_col = ws.max_column
    if max_col is None:
        # read_only モードでmax_columnがNoneになる場合、row1を直接走査
        for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
            if cell.column >= 5 and cell.value:
                cols.append(cell.column)
        return cols
    for c in range(5, max_col + 1):
        if ws.cell(1, c).value:
            cols.append(c)
    return cols


def _copy_cell(src, dst):
    import copy as _copy
    dst.value = src.value
    if src.has_style:
        dst.font        = _copy.copy(src.font)
        dst.border      = _copy.copy(src.border)
        dst.fill        = _copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection  = _copy.copy(src.protection)
        dst.alignment   = _copy.copy(src.alignment)


def _replace_col_refs(formula, col_map):
    """数式内の列参照を col_map に従って置換する。"""
    sorted_keys = sorted(col_map.keys(), key=len, reverse=True)
    pattern = re.compile(
        r'(\$?)(' + '|'.join(re.escape(k) for k in sorted_keys) + r')(\$?\d+)'
    )
    def replacer(m):
        return m.group(1) + col_map[m.group(2)] + m.group(3)
    return pattern.sub(replacer, formula)


def _copy_cell_with_formula(src, dst, col_map=None):
    """数式を保持しつつセルをコピー。col_map 指定時は列参照を置換。"""
    import copy as _copy
    v = src.value
    if col_map and isinstance(v, str) and v.startswith('='):
        v = _replace_col_refs(v, col_map)
    dst.value = v
    if src.has_style:
        dst.font          = _copy.copy(src.font)
        dst.border        = _copy.copy(src.border)
        dst.fill          = _copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.protection    = _copy.copy(src.protection)
        dst.alignment     = _copy.copy(src.alignment)


def _apply_print_settings(ws, row_breaks, scale=60):
    from openpyxl.worksheet.pagebreak import Break
    ws.sheet_view.view = 'pageBreakPreview'
    ws.sheet_view.zoomScale = 60
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.scale = scale
    ws.page_margins.left   = 0.3937007874015748
    ws.page_margins.right  = 0.0
    ws.page_margins.top    = 0.0
    ws.page_margins.bottom = 0.0
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = False
    for rn in sorted(set(row_breaks)):
        ws.row_breaks.append(Break(id=rn))


def remove_external_links(path):
    """xlsxファイルからexternalLinksをXMLレベルで除去して上書き保存"""
    import zipfile, tempfile, shutil

    dir_name = os.path.dirname(os.path.abspath(path))
    tmp_fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir=dir_name)
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(path, 'r') as zin:
            with zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if 'externalLink' in item.filename:
                        continue
                    data = zin.read(item.filename)
                    try:
                        text = data.decode('utf-8')
                        text = re.sub(r'<externalReferences>.*?</externalReferences>', '', text, flags=re.DOTALL)
                        text = re.sub(r'<externalReference[^/]*/>', '', text)
                        text = re.sub(r'<Override[^>]*externalLink[^>]*/>', '', text)
                        text = re.sub(r'<Relationship[^>]*externalLink[^>]*/>', '', text)
                        if 'workbook.xml' in item.filename and 'rels' not in item.filename:
                            seen = set()
                            def dedup(m):
                                key = m.group(2)
                                if key in seen:
                                    return ''
                                seen.add(key)
                                return m.group(0)
                            text = re.sub(
                                r'(<definedName[^>]*_xlnm\.Print_Titles[^>]*localSheetId="(\d+)"[^>]*>.*?</definedName>)',
                                dedup, text, flags=re.DOTALL
                            )
                        data = text.encode('utf-8')
                    except Exception:
                        pass
                    zout.writestr(item, data)

        backup = path + '.bak'
        shutil.copy2(path, backup)
        shutil.move(tmp, path)
        try:
            os.remove(backup)
        except Exception:
            pass
        return True

    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        return False


# ── Web用：発注書結合 ────────────────────────────────────────────

def web_merge_orders(file_paths):
    """
    複数の発注書Excelを結合して bytes を返す。
    コピーシート（「コピー」を含む）を優先して使用。
    3店舗ずつ1シートに配置（パソコン版 order_merge_build と同じ形式）。
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    today = date.today()
    mmdd = f"{today.month:02d}{today.day:02d}"
    STORES_PER_SHEET = 3

    def _get_best_sheet_name(wb, keyword):
        """keyword を含むシートを検索。コピーシートを優先。"""
        sheets = wb.sheetnames
        copy_sheets = [s for s in sheets if keyword in s and 'コピー' in s]
        if copy_sheets:
            return copy_sheets[0]
        normal_sheets = [s for s in sheets if keyword in s]
        return normal_sheets[0] if normal_sheets else None

    # ── 各ファイルから店舗スロットを収集 ──
    inaho_slots = []
    sankyo_slots = []

    for path in file_paths:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for slots, keyword in [(inaho_slots, '稲穂'), (sankyo_slots, '三共')]:
                sname = _get_best_sheet_name(wb, keyword)
                if sname is None:
                    continue
                ws = wb[sname]
                cols = hacchu_get_store_cols(ws)
                for col in cols:
                    store_name = str(ws.cell(1, col).value or '').strip()
                    slots.append({
                        'path': path,
                        'sheet_name': sname,
                        'col': col,
                        'store': store_name or os.path.basename(path),
                    })
            wb.close()
        except Exception as e:
            raise RuntimeError(f'ファイル読み込みエラー: {os.path.basename(path)}: {e}')

    if not inaho_slots and not sankyo_slots:
        raise RuntimeError('稲穂・三共シートの店舗列が見つかりません')

    # ── 結合ワークブック作成 ──
    wb_dst = openpyxl.Workbook()
    wb_dst.remove(wb_dst.active)

    def _build_kind(slots, kind_label, page_breaks, scale):
        if not slots:
            return
        n_groups = (len(slots) + STORES_PER_SHEET - 1) // STORES_PER_SHEET

        for g_idx in range(0, len(slots), STORES_PER_SHEET):
            group = slots[g_idx: g_idx + STORES_PER_SHEET]
            g_num = g_idx // STORES_PER_SHEET

            # シート名: グループ1つなら「【稲穂】」、複数なら「【稲穂】1」「【稲穂】2」
            sheet_name = kind_label if n_groups == 1 else f'{kind_label}{g_num + 1}'
            dst_ws = wb_dst.create_sheet(sheet_name)

            # A-D列を最初のスロットのファイルからコピー
            first = group[0]
            wb0 = openpyxl.load_workbook(first['path'], data_only=False)
            src0 = wb0[first['sheet_name']]

            fixed_cols = list(range(1, 5))
            for c in fixed_cols:
                sl = get_column_letter(c)
                if sl in src0.column_dimensions:
                    dst_ws.column_dimensions[sl].width = src0.column_dimensions[sl].width
            for rn, rd in src0.row_dimensions.items():
                dst_ws.row_dimensions[rn].height = rd.height
            for row in src0.iter_rows(max_row=src0.max_row):
                for sc in row:
                    if sc.column not in fixed_cols:
                        continue
                    dc = dst_ws.cell(row[0].row, sc.column)
                    _copy_cell_with_formula(sc, dc)
            dst_ws.cell(1, 3).value = int(mmdd)
            wb0.close()

            # 各スロットの店舗列をコピー
            next_dst_col = 5
            for slot in group:
                wb_add = openpyxl.load_workbook(slot['path'], data_only=False)
                src_add = wb_add[slot['sheet_name']]
                src_col = slot['col']
                dst_col = next_dst_col

                sl = get_column_letter(src_col)
                dl = get_column_letter(dst_col)
                if sl in src_add.column_dimensions:
                    dst_ws.column_dimensions[dl].width = src_add.column_dimensions[sl].width

                formula_col_map = {sl: dl}
                for row in src_add.iter_rows(max_row=src_add.max_row):
                    for sc in row:
                        if sc.column != src_col:
                            continue
                        dc = dst_ws.cell(row[0].row, dst_col)
                        _copy_cell_with_formula(sc, dc, formula_col_map)

                wb_add.close()
                next_dst_col += 1

            # ヘッダー行フォント・高さ調整
            store_cols_dst = list(range(5, next_dst_col))
            header_rows = [1, 2] + [b + 1 for b in page_breaks] + [b + 2 for b in page_breaks]
            for r in header_rows:
                for c in store_cols_dst:
                    cell = dst_ws.cell(r, c)
                    if cell.value is not None:
                        cell.alignment = Alignment(
                            wrap_text=True, shrink_to_fit=False,
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                        )
                        cell.font = Font(
                            name=cell.font.name, size=12, bold=cell.font.bold,
                            italic=cell.font.italic, color=cell.font.color,
                        )
                if r in [1] + [b + 1 for b in page_breaks]:
                    dst_ws.row_dimensions[r].height = 48
                else:
                    dst_ws.row_dimensions[r].height = 22.5

            _apply_print_settings(dst_ws, page_breaks, scale=scale)

    _build_kind(inaho_slots, '【稲穂】', [54, 108, 165, 219, 276], scale=67)
    _build_kind(sankyo_slots, '【三共】', [59], scale=61)

    # bytes として返す
    buf = io.BytesIO()
    wb_dst.save(buf)
    buf.seek(0)
    return buf.read()
